#!/usr/bin/env python3
"""Neftel Table S2 -> data/raw/neftel_signatures/neftel_metamodules.tsv.

Table S2 publishes SIX meta-modules: MES1, MES2, AC, OPC, NPC1, NPC2, plus two
cell-cycle programs (G1/S, G2/M). Neftel et al. collapse MES1+MES2 to "MES-like"
and NPC1+NPC2 to "NPC-like" for most analyses, and CLAUDE.md §2.1 specifies four
states. We do the same collapse and record it.

The cell-cycle columns are NOT states and are excluded.

This input is manual (see DEVIATIONS.md): PMC gates the download path with a
proof-of-work anti-bot challenge. The expected SHA256 is asserted here so a
substituted file fails loudly rather than silently changing the state definitions.
"""

from __future__ import annotations

import hashlib
from pathlib import Path

import openpyxl
import pandas as pd

REPO = Path(__file__).resolve().parent.parent
SRC = REPO / "data" / "raw" / "neftel_signatures" / "NIHMS1532254-supplement-9.xlsx"
OUT = REPO / "data" / "raw" / "neftel_signatures" / "neftel_metamodules.tsv"

EXPECTED_SHA256 = "208e73ab3d22c494caf85c867d69dc6be38df3fc62ab1f043d7fcc5441066277"

COLLAPSE = {"MES1": "MES", "MES2": "MES", "AC": "AC", "OPC": "OPC",
            "NPC1": "NPC", "NPC2": "NPC"}
CELL_CYCLE = {"G1/S", "G2/M"}


def main() -> int:
    assert SRC.exists(), f"missing manual input {SRC} (see DEVIATIONS.md)"
    got = hashlib.sha256(SRC.read_bytes()).hexdigest()
    assert got == EXPECTED_SHA256, (
        f"SHA256 mismatch for {SRC.name}\n  expected {EXPECTED_SHA256}\n"
        f"  got      {got}\nRefusing to redefine cell states from an unverified file.")
    print(f"source verified: {SRC.name}  sha256 ok")

    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    hdr_i = next(i for i, r in enumerate(rows)
                 if r and sum(1 for c in r if isinstance(c, str)
                              and c.strip() in set(COLLAPSE) | CELL_CYCLE) >= 4)
    header = [str(c).strip() if c is not None else "" for c in rows[hdr_i]]
    print(f"header row {hdr_i+1}: {[h for h in header if h]}")

    out = []
    for j, name in enumerate(header):
        if name in CELL_CYCLE:
            continue
        if name not in COLLAPSE:
            continue
        genes = []
        for r in rows[hdr_i + 1:]:
            if j < len(r) and isinstance(r[j], str) and r[j].strip():
                genes.append(r[j].strip())
        print(f"  {name:<5} -> {COLLAPSE[name]:<4} {len(genes):>3} genes")
        out += [{"state": COLLAPSE[name], "gene": g, "source_module": name}
                for g in genes]

    df = pd.DataFrame(out)
    print("\ncollapse applied (as Neftel et al. do; CLAUDE.md §2.1 uses four states):")
    for st, d in df.groupby("state"):
        mods = sorted(d["source_module"].unique())
        print(f"  {st:<4} <- {'+'.join(mods):<10} "
              f"{len(d):>3} rows, {d['gene'].nunique():>3} unique genes")

    ded = df.drop_duplicates(subset=["state", "gene"])[["state", "gene"]]
    ded.to_csv(OUT, sep="\t", index=False)
    print(f"\nwrote {OUT.relative_to(REPO)} ({len(ded)} state-gene pairs)")

    overlap = (df.groupby("gene")["state"].nunique() > 1).sum()
    print(f"genes appearing in more than one state: {overlap}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
